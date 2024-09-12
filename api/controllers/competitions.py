import logging
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from fastapi import HTTPException

from models.competitions import Competition, CompetitionResultsExport, CompetitionType, CompetitionExport, CompetitionResults
from models.results import RunResultsExport
from models.svg_data import SvgData
from models.pilots import GenderEnum
from controllers.utils import UtilsCtrl

log = logging.getLogger(__name__)

class CompCtrl:
    @staticmethod
    def start():
        Competition.createIndexes()

    @staticmethod
    def comp_to_xlsx(comp_results:CompetitionResultsExport, comp:Competition):
        ret = None
        wb = Workbook()
        ws = wb.active
        ws.title = f"overall {comp.type}"

        if comp.type == CompetitionType.solo:
            ws.cell(column=1, row=1, value="Rank")
            ws.cell(column=2, row=1, value="Number")
            ws.cell(column=3, row=1, value="First Name")
            ws.cell(column=4, row=1, value="Last Name")
            ws.cell(column=5, row=1, value="Nat")
            ws.cell(column=6, row=1, value="Gender")
            ws.cell(column=7, row=1, value="Glider")
            ws.cell(column=8, row=1, value="Sponsor")
            ws.cell(column=9, row=1, value="CIVL ID")
            ws.cell(column=10, row=1, value="Score")

        if comp.type == CompetitionType.synchro:
            ws.cell(column=1, row=1, value="Rank")
            ws.cell(column=2, row=1, value="Team")
            ws.cell(column=3, row=1, value="Number")
            ws.cell(column=4, row=1, value="First Name")
            ws.cell(column=5, row=1, value="Last Name")
            ws.cell(column=6, row=1, value="Nat")
            ws.cell(column=7, row=1, value="Gender")
            ws.cell(column=8, row=1, value="Glider")
            ws.cell(column=9, row=1, value="Sponsor")
            ws.cell(column=10, row=1, value="CIVL ID")
            ws.cell(column=11, row=1, value="Score")

        all_results = comp_results.results["overall"]
        if comp.type == CompetitionType.solo:
            awt_results = list(filter(lambda r: r.pilot.is_awt, comp_results.results["overall"]))
            awq_results = list(filter(lambda r: not r.pilot.is_awt, comp_results.results["overall"]))
            all_results = awt_results + awq_results

        rank = 0
        row = 1
        score = 100
        first_awq = False
        for res in all_results:
            rank += 1

            if any(re.search(r"^aw[tqs]", s) for s in comp.seasons):
                if comp.type == CompetitionType.solo and not first_awq and not res.pilot.is_awt:
                    score = 80
                    first_awq = True

            if comp.type == CompetitionType.solo:
                row += 1
                ws.cell(column=1, row=row, value=rank)
                ws.cell(column=2, row=row, value=res.pilot.civlid)
                ws.cell(column=3, row=row, value=res.pilot.name.split(" ", 1)[0])
                ws.cell(column=4, row=row, value=res.pilot.name.split(" ", 1)[1])
                ws.cell(column=5, row=row, value=res.pilot.country.upper())
                ws.cell(column=6, row=row, value="M" if res.pilot.gender == "man" else "F")
                ws.cell(column=7, row=row, value=f"")
                ws.cell(column=8, row=row, value=f"")
                ws.cell(column=9, row=row, value=res.pilot.civlid)
                ws.cell(column=10, row=row, value=score)

            if comp.type == CompetitionType.synchro:
                for i, pilot in enumerate(res.team.pilots):
                    row += 1
                    ws.cell(column=1, row=row, value=rank)
                    ws.cell(column=2, row=row, value=f"{res.team.name}")
                    ws.cell(column=3, row=row, value=pilot.civlid)
                    ws.cell(column=4, row=row, value=pilot.name.split(" ", 1)[0])
                    ws.cell(column=5, row=row, value=pilot.name.split(" ", 1)[1])
                    ws.cell(column=6, row=row, value=pilot.country.upper())
                    ws.cell(column=7, row=row, value=f"M")
                    ws.cell(column=8, row=row, value=f"")
                    ws.cell(column=9, row=row, value=f"")
                    ws.cell(column=10, row=row, value=pilot.civlid)
                    ws.cell(column=11, row=row, value=score)
                ws.merge_cells(start_row=row-1, start_column=1, end_row=row, end_column=1)
                ws.merge_cells(start_row=row-1, start_column=2, end_row=row, end_column=2)
                ws.merge_cells(start_row=row-1, start_column=11, end_row=row, end_column=11)

            score = score - 1

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as xlsx:
            ret = xlsx.name
            wb.save(filename=ret)

        return ret

    @staticmethod
    def run_to_xlsx(run:RunResultsExport, comp: Competition):
        if not run.final:
            raise HTTPException(400, f"Can't export run because it's not marked as final")
        ret = None
        wb = Workbook()
        ws = wb.active
        ws.title = f"run {comp.type}"

        if comp.type == CompetitionType.solo:
            ws.cell(column=1, row=1, value="Rank")
            ws.cell(column=2, row=1, value="Number")
            ws.cell(column=3, row=1, value="First Name")
            ws.cell(column=4, row=1, value="Last Name")
            ws.cell(column=5, row=1, value="Nat")
            ws.cell(column=6, row=1, value="Gender")
            ws.cell(column=7, row=1, value="Glider")
            ws.cell(column=8, row=1, value="Sponsor")
            ws.cell(column=9, row=1, value="CIVL ID")
            ws.cell(column=10, row=1, value="Score")

        if comp.type == CompetitionType.synchro:
            ws.cell(column=1, row=1, value="Rank")
            ws.cell(column=2, row=1, value="Team")
            ws.cell(column=3, row=1, value="Number")
            ws.cell(column=4, row=1, value="First Name")
            ws.cell(column=5, row=1, value="Last Name")
            ws.cell(column=6, row=1, value="Nat")
            ws.cell(column=7, row=1, value="Gender")
            ws.cell(column=8, row=1, value="Glider")
            ws.cell(column=9, row=1, value="Sponsor")
            ws.cell(column=10, row=1, value="CIVL ID")
            ws.cell(column=11, row=1, value="Score")


        run.results["overall"].sort(key=lambda e: -e.final_marks.score)
        all_results = run.results["overall"]

        if comp.type == CompetitionType.solo:
            awt_results = list(filter(lambda r: r.pilot.is_awt, run.results["overall"]))
            awq_results = list(filter(lambda r: not r.pilot.is_awt, run.results["overall"]))
            all_results = awt_results + awq_results

        rank = 0
        row = 1
        score = 100
        first_awq = False
        for res in all_results:
            rank += 1

            if any(re.search(r"^aw[tqs]", s) for s in comp.seasons):
                if comp.type == CompetitionType.solo and not first_awq and not res.pilot.is_awt:
                    score = 80
                    first_awq = True

            if comp.type == CompetitionType.solo:
                row += 1
                ws.cell(column=1, row=row, value=rank)
                ws.cell(column=2, row=row, value=res.pilot.civlid)
                ws.cell(column=3, row=row, value=res.pilot.name.split(" ", 1)[0])
                ws.cell(column=4, row=row, value=res.pilot.name.split(" ", 1)[1])
                ws.cell(column=5, row=row, value=res.pilot.country.upper())
                ws.cell(column=6, row=row, value="M" if res.pilot.gender == "man" else "F")
                ws.cell(column=7, row=row, value=f"")
                ws.cell(column=8, row=row, value=f"")
                ws.cell(column=9, row=row, value=res.pilot.civlid)
                ws.cell(column=10, row=row, value=score)

            if comp.type == CompetitionType.synchro:
                for i, pilot in enumerate(res.team.pilots):
                    row += 1
                    ws.cell(column=1, row=row, value=rank)
                    ws.cell(column=2, row=row, value=f"{res.team.name}")
                    ws.cell(column=3, row=row, value=pilot.civlid)
                    ws.cell(column=4, row=row, value=pilot.name.split(" ", 1)[0])
                    ws.cell(column=5, row=row, value=pilot.name.split(" ", 1)[1])
                    ws.cell(column=6, row=row, value=pilot.country.upper())
                    ws.cell(column=7, row=row, value=f"M")
                    ws.cell(column=8, row=row, value=f"")
                    ws.cell(column=9, row=row, value=f"")
                    ws.cell(column=10, row=row, value=pilot.civlid)
                    ws.cell(column=11, row=row, value=score)
                ws.merge_cells(start_row=row-1, start_column=1, end_row=row, end_column=1)
                ws.merge_cells(start_row=row-1, start_column=2, end_row=row, end_column=2)
                ws.merge_cells(start_row=row-1, start_column=11, end_row=row, end_column=11)

            score = score - 1

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as xlsx:
            ret = xlsx.name
            wb.save(filename=ret)

        return ret

    @staticmethod
    def svg_overall(competition: CompetitionResults, result_type: str = 'overall', animated: int = -1):

        competition.results["awt"] = []
        competition.results["awq"] = []
        competition.results["women"] = []

        for result in competition.results["overall"]:
            if result.pilot.is_awt:
                competition.results["awt"].append(result)
            else:
                competition.results["awq"].append(result)

            if result.pilot.gender == GenderEnum.woman:
                competition.results["women"].append(result)

        if result_type not in competition.results:
            raise HTTPException(status_code=404, detail=f"Result Type not found for this competition")
        results = []
        for rank, result in enumerate(competition.results[result_type]):
            results.append(SvgData(
                rank=rank+1,
                country=result.pilot.country if competition.type == "solo" else None,
                name=result.pilot.name if competition.type == "solo" else result.team.name,
                score="%.3f" % result.score
            ))

        return UtilsCtrl.svg(results, animated=animated)

    @staticmethod
    def svg_run(competition: CompetitionResults, run: int, result_type: str = 'overall', animated: int = -1):
        try:
            competition = competition.runs_results[run-1]
        except:
            raise HTTPException(status_code=404, detail=f"Run not found for this competition")

        competition.results["awt"] = []
        competition.results["awq"] = []
        competition.results["women"] = []

        for result in competition.results["overall"]:
            if result.pilot.is_awt:
                competition.results["awt"].append(result)
            else:
                competition.results["awq"].append(result)

            if result.pilot.gender == GenderEnum.woman:
                competition.results["women"].append(result)

        if result_type not in competition.results:
            raise HTTPException(status_code=404, detail=f"Result Type not found for this run")
        results = []
        competition.results[result_type].sort(key=lambda e: -e.final_marks.score)
        for rank, result in enumerate(competition.results[result_type]):
            results.append(SvgData(
                rank=rank+1,
                country=result.pilot.country if competition.type == "solo" else None,
                name=result.pilot.name if competition.type == "solo" else result.team.name,
                score="%.3f" % result.final_marks.score
            ))

        return UtilsCtrl.svg(results, animated=animated)
